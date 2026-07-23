import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from app.models import Trabajo

def generar_contenido_txt(trabajo: Trabajo) -> str:
    """
    Genera la cadena de texto tabulada para el optimizador .txt
    respetando el formato exacto de la aplicación.
    """
    cortes_txt = ""
    
    for c in trabajo.cortes:
        cant = c.cantidad
        largo = c.largo_mm
        ancho = c.ancho_mm
        descripcion = c.descripcion
        tl = c.tapacanto_largo
        ta = c.tapacanto_ancho
        
        # Leemos el flag de rotación de cada corte (1 o 0)
        rota_val = 1 if c.puede_rotar else 0

        # Lógica de conversión de tapacantos a lados
        if tl == 1:
            canto_arr, canto_aba = 1, 0
        elif tl == 2:
            canto_arr, canto_aba = 1, 1
        elif tl == 3:
            canto_arr, canto_aba = 2, 0
        elif tl == 4:
            canto_arr, canto_aba = 2, 2
        else:
            canto_arr, canto_aba = 0, 0

        if ta == 1:
            canto_izq, canto_der = 1, 0
        elif ta == 2:
            canto_izq, canto_der = 1, 1
        elif ta == 3:
            canto_izq, canto_der = 2, 0
        elif ta == 4:
            canto_izq, canto_der = 2, 2
        else:
            canto_izq, canto_der = 0, 0

        # Construcción de la cadena de texto tabulada exacta
        cortes_txt += f"{cant}\t{largo}\t{ancho}\t{descripcion}\t{rota_val}\t"
        cortes_txt += f"{canto_arr}\t{canto_aba}\t{canto_izq}\t{canto_der}\n"

    return cortes_txt

def generar_excel(trabajo: Trabajo) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) con los datos de cortes y formato ajustado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cortes"
    
    # Encabezados
    headers = [
        "Cantidad", 
        "Base (Largo)", 
        "Altura (Ancho)", 
        "Detalle / Descripción", 
        "Rotar", 
        "Canto Arriba", 
        "Canto Abajo", 
        "Canto Izquierda", 
        "Canto Derecha"
    ]
    ws.append(headers)
    
    # Dar formato a las cabeceras (verde emerald)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for c in trabajo.cortes:
        cant = c.cantidad
        largo = c.largo_mm
        ancho = c.ancho_mm
        descripcion = c.descripcion
        tl = c.tapacanto_largo
        ta = c.tapacanto_ancho
        
        rota_val = 1 if c.puede_rotar else 0

        if tl == 1: canto_arr, canto_aba = 1, 0
        elif tl == 2: canto_arr, canto_aba = 1, 1
        elif tl == 3: canto_arr, canto_aba = 2, 0
        elif tl == 4: canto_arr, canto_aba = 2, 2
        else: canto_arr, canto_aba = 0, 0

        if ta == 1: canto_izq, canto_der = 1, 0
        elif ta == 2: canto_izq, canto_der = 1, 1
        elif ta == 3: canto_izq, canto_der = 2, 0
        elif ta == 4: canto_izq, canto_der = 2, 2
        else: canto_izq, canto_der = 0, 0
        
        ws.append([cant, largo, ancho, descripcion, rota_val, canto_arr, canto_aba, canto_izq, canto_der])
    
    # Ajustar ancho de columnas automáticamente
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        # Añadimos un poco de padding
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
